import os
import secrets
import string
import uuid
from datetime import datetime, timedelta
from dataclasses import dataclass, field
from typing import List, Optional, Tuple

from openpyxl import Workbook, load_workbook

from models import Order, Customer, User, Inventory, ReturnRequest, ReturnStatus
from enums import OrderStatus, CustomerType, UserRole, InventoryStatus, ReturnReason
from database import OrderRepository, CustomerRepository, InventoryRepository, ReturnRequestRepository
from database.user_repository import UserRepository, UserAlreadyExistsError
from database.customer_repository import CustomerAlreadyExistsError
from database.inventory_repository import InventoryNotFoundError
from database.return_request_repository import ReturnRequestAlreadyExistsError


REQUIRED_HEADERS = [
    "客户类型", "会员名", "销售员", "订单号", "运单号", "订单状态",
    "下单时间", "付款时间", "订单最晚发货日期", "购买产品", "购买数量", "退货处理号"
]

INVENTORY_HEADERS = [
    "产品类型",
    "品牌",
    "产品名称",
    "产品型号",
    "产品ID",
    "库存数量",
    "已售数量",
    "状态",
    "预计补货时间",
]


@dataclass
class CreatedAccountInfo:
    company_name: str = ""
    username: str = ""
    password: str = ""
    display_name: str = ""
    role: str = ""
    customer_id: str = ""


@dataclass
class ImportResult:
    orders_created: int = 0
    customers_created: int = 0
    users_created: int = 0
    return_requests_created: int = 0
    return_requests_skipped: int = 0
    created_accounts: List[CreatedAccountInfo] = field(default_factory=list)
    errors: List[str] = field(default_factory=list)


class ExcelService:
    def __init__(self):
        self._order_repo = OrderRepository()
        self._customer_repo = CustomerRepository()
        self._user_repo = UserRepository()
        self._inventory_repo = InventoryRepository()
        self._return_request_repo = ReturnRequestRepository()

    @staticmethod
    def _generate_password(length: int = 10) -> str:
        alphabet = string.ascii_letters + string.digits + "!@#$%&*"
        password = [
            secrets.choice(string.ascii_lowercase),
            secrets.choice(string.ascii_uppercase),
            secrets.choice(string.digits),
            secrets.choice("!@#$%&*"),
        ]
        for _ in range(length - 4):
            password.append(secrets.choice(alphabet))
        secrets.SystemRandom().shuffle(password)
        return ''.join(password)

    @staticmethod
    def _generate_username(name: str) -> str:
        chars = list(name)
        if len(chars) > 8:
            prefix = ''.join(chars[:8])
        elif len(chars) >= 3:
            prefix = ''.join(chars)
        else:
            prefix = "customer"
        
        suffix = secrets.token_hex(2)
        return f"{prefix}_{suffix}"

    @staticmethod
    def _parse_datetime(s: str) -> Optional[datetime]:
        s = s.strip()
        if not s:
            return None
        
        formats = [
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d",
            "%Y/%m/%d %H:%M:%S",
            "%Y/%m/%d",
            "%m/%d/%Y %H:%M:%S",
            "%m/%d/%Y",
            "%m/%d/%y %H:%M:%S",
            "%m/%d/%y",
        ]
        
        for fmt in formats:
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                continue
        
        return None

    def parse_excel(self, file_path: str) -> Tuple[List[Order], List[str]]:
        orders = []
        errors = []
        
        try:
            wb = load_workbook(file_path)
        except Exception as e:
            return [], [f"Failed to open Excel file: {e}"]
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            headers = {}
            for col_idx, cell in enumerate(sheet[1]):
                if cell.value:
                    headers[cell.value] = col_idx
            
            missing_headers = [h for h in REQUIRED_HEADERS if h not in headers]
            if missing_headers:
                errors.append(f"Sheet '{sheet_name}': Missing headers: {missing_headers}")
                continue
            
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                try:
                    def get_value(header: str) -> str:
                        col = headers.get(header)
                        if col is not None and col < len(row):
                            val = row[col].value
                            return str(val) if val is not None else ""
                        return ""

                    order_time_str = get_value("下单时间")
                    order_time = self._parse_datetime(order_time_str)
                    if not order_time:
                        errors.append(f"Row {row_idx}: Invalid order time: {order_time_str}")
                        continue

                    payment_time_str = get_value("付款时间")
                    payment_time = self._parse_datetime(payment_time_str)

                    ship_deadline_str = get_value("订单最晚发货日期")
                    ship_deadline = self._parse_datetime(ship_deadline_str)
                    quantity_str = get_value("购买数量")
                    try:
                        quantity = int(quantity_str)
                    except ValueError:
                        quantity = 1
                    
                    order = Order(
                        customer_type=int(CustomerType.from_string(get_value("客户类型"))),
                        customer_name=get_value("会员名"),
                        sales=get_value("销售员"),
                        order_id=get_value("订单号"),
                        tracking_number=get_value("运单号"),
                        status=int(OrderStatus.from_string(get_value("订单状态"))),
                        order_time=order_time,
                        payment_time=payment_time,
                        ship_deadline=ship_deadline,
                        product_id=get_value("购买产品"),
                        quantity=quantity,
                        return_request_id=get_value("退货处理号"),
                    )
                    
                    if not order.check_entity():
                        errors.append(f"Row {row_idx}: Invalid order data")
                        continue
                    
                    try:
                        self._inventory_repo.get_inventory_by_id(order.product_id)
                    except InventoryNotFoundError:
                        errors.append(f"Row {row_idx}: Product ID '{order.product_id}' not found in inventory")
                        continue
                    
                    orders.append(order)
                    
                except Exception as e:
                    errors.append(f"Row {row_idx}: Error parsing: {e}")
        
        wb.close()
        return orders, errors

    def import_excel_with_customers_and_users(
        self, file_path: str
    ) -> ImportResult:
        result = ImportResult()
        
        orders, parse_errors = self.parse_excel(file_path)
        result.errors.extend(parse_errors)
        
        if not orders:
            return result
        
        company_map = {}
        customer_id_map = {}
        sales_map = set()
        
        for order in orders:
            if order.customer_name:
                company_map[order.customer_name] = order.customer_type
            if order.sales:
                sales_map.add(order.sales)
        
        for company_name, customer_type in company_map.items():
            try:
                customer, is_new = self._get_or_create_customer(
                    company_name, customer_type
                )
                customer_id_map[company_name] = customer.customer_id
                
                if is_new:
                    result.customers_created += 1
            except Exception as e:
                result.errors.append(f"Failed to create customer '{company_name}': {e}")
        
        for sales_name in sales_map:
            try:
                account_info = self._create_sales_user(sales_name)
                if account_info:
                    result.created_accounts.append(account_info)
                    result.users_created += 1
            except UserAlreadyExistsError:
                pass
            except Exception as e:
                result.errors.append(f"Failed to create sales user '{sales_name}': {e}")
        
        for order in orders:
            try:
                if order.customer_name in customer_id_map:
                    order.customer_id = customer_id_map[order.customer_name]
                
                self._order_repo.create_order(order)
                result.orders_created += 1

                try:
                    self._inventory_repo.update_stock(order.product_id, -order.quantity)
                except InventoryNotFoundError:
                    result.errors.append(
                        f"Inventory not found for product '{order.product_id}' in order '{order.order_id}'"
                    )
                except Exception as inv_err:
                    result.errors.append(
                        f"Failed to update inventory for order '{order.order_id}': {inv_err}"
                    )
                
                if order.status in OrderStatus.get_return_statuses():
                    try:
                        return_request = ReturnRequest(
                            order_id=order.order_id,
                            product_id=order.product_id,
                            quantity=order.quantity,
                            reason=int(ReturnReason.OTHER),
                            customer_name=order.customer_name,
                            status=int(ReturnStatus.PENDING),
                        )
                        if order.return_request_id:
                            return_request.return_request_id = order.return_request_id
                        
                        self._return_request_repo.create_return_request(return_request)
                        
                        if not order.return_request_id:
                            order.return_request_id = return_request.return_request_id
                            self._order_repo.update_order(order)
                        
                        result.return_requests_created += 1
                    except ReturnRequestAlreadyExistsError:
                        result.return_requests_skipped += 1
                    except Exception as ret_err:
                        result.errors.append(
                            f"Failed to create return request for order '{order.order_id}': {ret_err}"
                        )
            except Exception as e:
                result.errors.append(f"Failed to create order '{order.order_id}': {e}")
        
        return result

    def _get_or_create_customer(
        self, company_name: str, customer_type: CustomerType
    ) -> Tuple[Customer, bool]:
        try:
            customer = self._customer_repo.get_customer_by_company_name(company_name)
            return customer, False
        except Exception:
            pass
        
        customer = Customer(
            customer_id=str(uuid.uuid4()),
            company_name=company_name,
            customer_type=int(customer_type),
            is_active=True,
        )
        
        try:
            self._customer_repo.create_customer(customer)
            return customer, True
        except CustomerAlreadyExistsError:
            customer = self._customer_repo.get_customer_by_company_name(company_name)
            return customer, False

    def _create_distributor_user(
        self, company_name: str, customer_id: str
    ) -> Optional[CreatedAccountInfo]:
        username = self._generate_username(company_name)
        password = self._generate_password()
        
        user = User(
            user_id=str(uuid.uuid4()),
            username=username,
            display_name=company_name,
            role=int(UserRole.VIEWER),
            is_active=True,
        )
        user.set_password(password)
        
        try:
            self._user_repo.create_user(user)
            return CreatedAccountInfo(
                company_name=company_name,
                username=username,
                password=password,
                display_name=company_name,
                role=str(UserRole.VIEWER),
                customer_id=customer_id,
            )
        except UserAlreadyExistsError:
            return None

    def _create_sales_user(self, sales_name: str) -> Optional[CreatedAccountInfo]:
        username = self._generate_username(sales_name)
        password = self._generate_password()
        
        user = User(
            user_id=str(uuid.uuid4()),
            username=username,
            display_name=sales_name,
            role=int(UserRole.OPERATOR),
            department="销售部",
            is_active=True,
        )
        user.set_password(password)
        
        self._user_repo.create_user(user)
        return CreatedAccountInfo(
            company_name="",
            username=username,
            password=password,
            display_name=sales_name,
            role=str(UserRole.OPERATOR),
            customer_id="",
        )

    @staticmethod
    def export_to_excel(
        headers: List[str],
        data: List[List[str]],
        custom_name: str = ""
    ) -> str:
        wb = Workbook()
        ws = wb.active
        
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)
        
        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, cell_value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=cell_value)
        
        export_dir = "./ExportFiles"
        if not os.path.exists(export_dir):
            os.makedirs(export_dir)
        
        timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        if custom_name:
            filename = f"{export_dir}/{custom_name}{timestamp}.xlsx"
        else:
            filename = f"{export_dir}/Export-{timestamp}.xlsx"
        
        wb.save(filename)
        wb.close()
        
        return filename

    @staticmethod
    def export_created_accounts_to_excel(
        accounts: List[CreatedAccountInfo]
    ) -> str:
        headers = ["用户名", "初始密码", "显示名称", "角色"]
        
        data = [
            [
                acc.username,
                acc.password,
                acc.display_name,
                acc.role,
            ]
            for acc in accounts
        ]
        
        return ExcelService.export_to_excel(headers, data, "CustomerAccounts")

    def parse_inventory_excel(self, file_path: str) -> Tuple[List[Inventory], List[str]]:
        inventory_items = []
        errors = []
        
        try:
            wb = load_workbook(file_path)
        except Exception as e:
            return [], [f"Failed to open Excel file: {e}"]
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            headers = {}
            for col_idx, cell in enumerate(sheet[1]):
                if cell.value:
                    headers[cell.value] = col_idx
            
            missing_headers = [h for h in INVENTORY_HEADERS if h not in headers]
            if missing_headers:
                errors.append(f"Sheet '{sheet_name}': Missing headers: {missing_headers}")
                continue
            
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                try:
                    def get_value(header: str) -> str:
                        col = headers.get(header)
                        if col is not None and col < len(row):
                            val = row[col].value
                            return str(val) if val is not None else ""
                        return ""
                    
                    stock_quantity_str = get_value("库存数量")
                    try:
                        stock_quantity = int(stock_quantity_str)
                    except ValueError:
                        stock_quantity = 0
                    
                    sold_quantity_str = get_value("已售数量")
                    try:
                        sold_quantity = int(sold_quantity_str)
                    except ValueError:
                        sold_quantity = 0
                    
                    expected_arrival_str = get_value("预计补货时间")
                    expected_arrival = self._parse_datetime(expected_arrival_str)
                    
                    status_str = get_value("状态")
                    status = InventoryStatus.from_string(status_str)
                    
                    product_id = get_value("产品ID")
                    if not product_id:
                        product_id = str(uuid.uuid4())
                    
                    inventory = Inventory(
                        product_id=product_id,
                        product_type=get_value("产品类型"),
                        manufacturer=get_value("品牌"),
                        product_name=get_value("产品名称"),
                        product_model=get_value("产品型号"),
                        stock_quantity=stock_quantity,
                        sold_quantity=sold_quantity,
                        status=int(status),
                        expected_arrival=expected_arrival,
                    )
                    
                    if not inventory.validate():
                        errors.append(f"Row {row_idx}: Invalid inventory data")
                        continue
                    
                    inventory_items.append(inventory)
                    
                except Exception as e:
                    errors.append(f"Row {row_idx}: Error parsing: {e}")
        
        wb.close()
        return inventory_items, errors

    def import_inventory_from_excel(self, file_path: str) -> Tuple[int, List[str]]:
        inventory_items, parse_errors = self.parse_inventory_excel(file_path)
        errors = list(parse_errors)
        
        if not inventory_items:
            return 0, errors
        
        success_count = 0
        for inventory in inventory_items:
            try:
                try:
                    existing = self._inventory_repo.get_inventory_by_id(inventory.product_id)
                    existing.product_type = inventory.product_type
                    existing.manufacturer = inventory.manufacturer
                    existing.product_name = inventory.product_name
                    existing.product_model = inventory.product_model
                    existing.stock_quantity = inventory.stock_quantity
                    existing.sold_quantity = inventory.sold_quantity
                    existing.status = inventory.status
                    existing.expected_arrival = inventory.expected_arrival
                    self._inventory_repo.update_inventory(existing)
                except InventoryNotFoundError:
                    self._inventory_repo.create_inventory(inventory)
                
                success_count += 1
            except Exception as e:
                errors.append(f"Failed to import inventory '{inventory.product_name}': {e}")
        
        return success_count, errors
